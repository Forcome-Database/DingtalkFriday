"""
Excel export service.

Generates .xlsx files using openpyxl with the same data structure
as the monthly summary table displayed in the frontend.
"""

import io
import logging
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.leave import get_monthly_summary

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
SUMMARY_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
SUMMARY_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
DATA_FONT = Font(name="Microsoft YaHei", size=11)
TOTAL_COL_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="E4E4E7"),
    right=Side(style="thin", color="E4E4E7"),
    top=Side(style="thin", color="E4E4E7"),
    bottom=Side(style="thin", color="E4E4E7"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


async def export_leave_data(
    year: int,
    dept_id: Optional[int] = None,
    leave_types: Optional[List[str]] = None,
    employee_name: Optional[str] = None,
    unit: str = "day",
) -> io.BytesIO:
    """
    Generate an Excel file containing the monthly leave summary.

    The export includes ALL matching rows (no pagination) plus
    a summary (totals) row at the bottom.

    Returns a BytesIO object containing the .xlsx data.
    """
    # Fetch all data without pagination
    data = await get_monthly_summary(
        year=year,
        dept_id=dept_id,
        leave_types=leave_types,
        employee_name=employee_name,
        unit=unit,
        page=1,
        page_size=999999,  # effectively no limit
        sort_by="name",
        sort_order="asc",
    )

    rows = data["list"]
    summary = data["summary"]
    unit_label = "天" if unit == "day" else "小时"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年请假数据"

    # ---- Header row ----
    headers = ["员工姓名", "部门"]
    for m in range(1, 13):
        headers.append(f"{m}月")
    headers.append(f"合计({unit_label})")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ---- Data rows ----
    for row_idx, row_data in enumerate(rows, start=2):
        # Name
        cell = ws.cell(row=row_idx, column=1, value=row_data["name"])
        cell.font = DATA_FONT
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER

        # Department
        cell = ws.cell(row=row_idx, column=2, value=row_data["dept"])
        cell.font = DATA_FONT
        cell.alignment = LEFT_ALIGN
        cell.border = THIN_BORDER

        # Monthly values
        for m_idx, val in enumerate(row_data["months"]):
            cell = ws.cell(row=row_idx, column=3 + m_idx, value=val if val else None)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            cell.number_format = "0.0"

        # Total column (highlighted)
        cell = ws.cell(row=row_idx, column=15, value=row_data["total"])
        cell.font = Font(name="Microsoft YaHei", bold=True, size=11)
        cell.fill = TOTAL_COL_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.number_format = "0.0"

    # ---- Summary row ----
    summary_row = len(rows) + 2
    cell = ws.cell(row=summary_row, column=1, value="合计")
    cell.fill = SUMMARY_FILL
    cell.font = SUMMARY_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    cell = ws.cell(row=summary_row, column=2, value=f"{summary['personCount']}人")
    cell.fill = SUMMARY_FILL
    cell.font = SUMMARY_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    for m_idx, val in enumerate(summary["months"]):
        cell = ws.cell(row=summary_row, column=3 + m_idx, value=val if val else None)
        cell.fill = SUMMARY_FILL
        cell.font = SUMMARY_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.number_format = "0.0"

    cell = ws.cell(row=summary_row, column=15, value=summary["total"])
    cell.fill = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")
    cell.font = SUMMARY_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    cell.number_format = "0.0"

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 14  # Name
    ws.column_dimensions["B"].width = 14  # Department
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10  # Month cols
    ws.column_dimensions["O"].width = 12  # Total column

    # ---- Freeze panes (freeze header + first 2 columns) ----
    ws.freeze_panes = "C2"

    # ---- Write to BytesIO ----
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(
        "Exported Excel: year=%d, dept_id=%s, rows=%d",
        year, dept_id, len(rows),
    )
    return output


def export_leave_detail(data: dict) -> io.BytesIO:
    """
    Generate an Excel file from leave detail records (single-date view).

    Columns: 姓名, 部门, 请假类型, 时间段, 时长, 状态
    """
    records = data.get("records", [])
    date_str = data.get("date", "")

    wb = Workbook()
    ws = wb.active
    ws.title = f"请假详情 {date_str}"

    headers = ["姓名", "部门", "请假类型", "时间段", "时长", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, rec in enumerate(records, 2):
        values = [
            rec.get("name", ""),
            rec.get("deptName", ""),
            rec.get("leaveType", ""),
            rec.get("timeDisplay", ""),
            rec.get("durationDisplay", ""),
            rec.get("status", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN if col > 2 else LEFT_ALIGN
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Exported leave detail: date=%s, records=%d", date_str, len(records))
    return output


def export_trip_detail(data: dict) -> io.BytesIO:
    """
    Generate an Excel file from trip detail records (single-date view).

    Columns: 姓名, 部门, 类型, 时间段, 时长
    """
    records = data.get("list", [])
    date_str = data.get("date", "")

    wb = Workbook()
    ws = wb.active
    ws.title = f"外出出差详情 {date_str}"

    headers = ["姓名", "部门", "类型", "时间段", "时长"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, rec in enumerate(records, 2):
        values = [
            rec.get("employeeName", ""),
            rec.get("deptName", ""),
            rec.get("tagName", ""),
            f"{rec.get('beginTime', '')} ~ {rec.get('endTime', '')}",
            f"{rec.get('durationHours', 0)}小时",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN if col > 2 else LEFT_ALIGN
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Exported trip detail: date=%s, records=%d", date_str, len(records))
    return output


async def export_trip_excel(
    year: int,
    dept_id: Optional[int] = None,
    trip_type: Optional[str] = None,
    employee_name: Optional[str] = None,
) -> io.BytesIO:
    """
    Export trip (出差/外出) data as an Excel file.

    The export includes ALL matching rows (no pagination) plus
    a summary (totals) row at the bottom.

    Columns: 姓名, 部门, 出差(天), 外出(天), 1月~12月, 合计(天) — 17 columns total.

    Returns a BytesIO object containing the .xlsx data.
    """
    from app.services.trip import get_trip_monthly_summary

    data = await get_trip_monthly_summary(
        year=year,
        dept_id=dept_id,
        trip_type=trip_type,
        employee_name=employee_name,
        page=1,
        page_size=999999,
    )

    rows = data["list"]
    summary = data["summary"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年外出出差统计"

    # ---- Header row ----
    headers = ["姓名", "部门", "出差(天)", "外出(天)"]
    headers += [f"{m}月" for m in range(1, 13)]
    headers.append("合计(天)")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ---- Data rows ----
    for row_idx, row in enumerate(rows, 2):
        values = [
            row["employeeName"],
            row.get("deptName", ""),
            row["tripDays"],
            row["outingDays"],
        ]
        for m in range(1, 13):
            values.append(row["months"].get(str(m), 0))
        values.append(row["totalDays"])

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    # ---- Summary row ----
    sum_row = len(rows) + 2
    sum_values = [
        "合计", "",
        summary["tripDays"],
        summary["outingDays"],
    ]
    for m in range(1, 13):
        sum_values.append(summary["months"].get(str(m), 0))
    sum_values.append(summary["totalDays"])

    for col, val in enumerate(sum_values, 1):
        cell = ws.cell(row=sum_row, column=col, value=val)
        cell.fill = SUMMARY_FILL
        cell.font = SUMMARY_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    for i in range(5, 17):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.column_dimensions[get_column_letter(17)].width = 10

    # ---- Freeze panes (freeze header + first 2 columns) ----
    ws.freeze_panes = "C2"

    # ---- Write to BytesIO ----
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(
        "Exported trip Excel: year=%d, dept_id=%s, rows=%d",
        year, dept_id, len(rows),
    )
    return output
